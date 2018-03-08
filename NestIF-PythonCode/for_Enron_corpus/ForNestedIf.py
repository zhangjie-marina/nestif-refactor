import os
from ete2 import Tree
from openpyxl import load_workbook

from collections import Counter


# ========================================================================
# Description: Tokenise an Excel formula using an implementation of
#              E. W. Bachtal's algorithm, found here:
#
#                  http://ewbi.blogs.com/develops/2004/12/excel_formula_p.html
#
#              Tested with Python v2.5 (win32)
#      Author: Robin Macharg
#   Copyright: Algorithm (c) E. W. Bachtal, this implementation (c) R. Macharg
#
# CVS Info:
# $Header: T:\\cvsarchive/Excel\040export\040&\040import\040XML/ExcelXMLTransform/EWBI_Javascript_port/jsport.py,v 1.5 2006/12/07 13:41:08 rmacharg Exp $
#
# Modification History
#
# Date         Author Comment
# =======================================================================
# 2006/11/29 - RMM  - Made strictly class-based.
#                     Added parse, render and pretty print methods
# 2006/11    - RMM  - RMM = Robin Macharg
#                           Created
# ========================================================================

# ========================================================================
#       Class: ExcelParserTokens
# Description: Inheritable container for token definitions
#
#  Attributes: Self explanatory
#
#     Methods: None
# ========================================================================



class ExcelParserTokens:
    TOK_TYPE_NOOP = "noop";
    TOK_TYPE_OPERAND = "operand";
    TOK_TYPE_FUNCTION = "function";
    TOK_TYPE_SUBEXPR = "subexpression";
    TOK_TYPE_ARGUMENT = "argument";
    TOK_TYPE_OP_PRE = "operator-prefix";
    TOK_TYPE_OP_IN = "operator-infix";
    TOK_TYPE_OP_POST = "operator-postfix";
    TOK_TYPE_WSPACE = "white-space";
    TOK_TYPE_UNKNOWN = "unknown"

    TOK_SUBTYPE_START = "start";
    TOK_SUBTYPE_STOP = "stop";
    TOK_SUBTYPE_TEXT = "text";
    TOK_SUBTYPE_NUMBER = "number";
    TOK_SUBTYPE_LOGICAL = "logical";
    TOK_SUBTYPE_ERROR = "error";
    TOK_SUBTYPE_RANGE = "range";
    TOK_SUBTYPE_MATH = "math";
    TOK_SUBTYPE_CONCAT = "concatenate";
    TOK_SUBTYPE_INTERSECT = "intersect";
    TOK_SUBTYPE_UNION = "union";


# ========================================================================
#       Class: f_token
# Description: Encapsulate a formula token
#
#  Attributes:   tvalue -
#                 ttype - See token definitions, above, for values
#              tsubtype - See token definitions, above, for values
#
#     Methods: f_token  - __init__()
# ========================================================================
class f_token:
    def __init__(self, value, type, subtype):
        self.tvalue = value
        self.ttype = type
        self.tsubtype = subtype


# ========================================================================
#       Class: f_tokens
# Description: An ordered list of tokens

#  Attributes:        items - Ordered list
#                     index - Current position in the list
#
#     Methods: f_tokens     - __init__()
#              f_token      - add()      - Add a token to the end of the list
#              None         - addRef()   - Add a token to the end of the list
#              None         - reset()    - reset the index to -1
#              Boolean      - BOF()      - End of list?
#              Boolean      - EOF()      - Beginning of list?
#              Boolean      - moveNext() - Move the index along one
#              f_token/None - current()  - Return the current token
#              f_token/None - next()     - Return the next token (leave the index unchanged)
#              f_token/None - previous() - Return the previous token (leave the index unchanged)
# ========================================================================
class f_tokens:
    def __init__(self):
        self.items = []
        self.index = -1

    def add(self, value, type, subtype=""):
        if (not subtype):
            subtype = ""
        token = f_token(value, type, subtype)
        self.addRef(token)
        return token

    def addRef(self, token):
        self.items.append(token)

    def reset(self):
        self.index = -1

    def BOF(self):
        return self.index <= 0

    def EOF(self):
        return self.index >= (len(self.items) - 1)

    def moveNext(self):
        if self.EOF():
            return False
        self.index += 1
        return True

    def current(self):
        if self.index == -1:
            return None
        return self.items[self.index]

    def next(self):
        if self.EOF():
            return None
        return self.items[self.index + 1]

    def previous(self):
        if self.index < 1:
            return None
        return self.items[self.index - 1]


# ========================================================================
#       Class: f_tokenStack
#    Inherits: ExcelParserTokens - a list of token values
# Description: A LIFO stack of tokens
#
#  Attributes:        items - Ordered list
#
#     Methods: f_tokenStack - __init__()
#              None         - push(token) - Push a token onto the stack
#              f_token/None - pop()       - Pop a token off the stack
#              f_token/None - token()     - Non-destructively return the top item on the stack
#              String       - type()      - Return the top token's type
#              String       - subtype()   - Return the top token's subtype
#              String       - value()     - Return the top token's value
# ========================================================================
class f_tokenStack(ExcelParserTokens):
    def __init__(self):
        self.items = []

    def push(self, token):
        self.items.append(token)

    def pop(self):
        token = self.items.pop()
        return f_token("", token.ttype, self.TOK_SUBTYPE_STOP)

    def token(self):
        # Note: this uses Pythons and/or "hack" to emulate C's ternary operator (i.e. cond ? exp1 : exp2)
        return ((len(self.items) > 0) and [self.items[len(self.items) - 1]] or [None])[0]

    def value(self):
        return ((self.token()) and [(self.token()).tvalue] or [""])[0]

    def type(self):
        t = self.token()
        return ((self.token()) and [(self.token()).ttype] or [""])[0]

    def subtype(self):
        return ((self.token()) and [(self.token()).tsubtype] or [""])[0]


# ========================================================================
#       Class: ExcelParser
# Description: Parse an Excel formula into a stream of tokens

#  Attributes:
#
#     Methods: f_tokens - getTokens(formula) - return a token stream (list)
# ========================================================================
class ExcelParser(ExcelParserTokens):
    def getTokens(self, formula):

        def currentChar():
            return formula[offset]

        def doubleChar():
            return formula[offset:offset + 2]

        def nextChar():
            # JavaScript returns an empty string if the index is out of bounds,
            # Python throws an IndexError.  We mimic this behaviour here.
            try:
                formula[offset + 1]
            except IndexError:
                return ""
            else:
                return formula[offset + 1]

        def EOF():
            return offset >= len(formula)

        tokens = f_tokens()
        tokenStack = f_tokenStack()
        offset = 0
        token = ""
        inString = False
        inPath = False
        inRange = False
        inError = False

        while (len(formula) > 0):
            if (formula[0] == " "):
                formula = formula[1:]
            else:
                if (formula[0] == "="):
                    formula = formula[1:]
                break;

                # state-dependent character evaluation (order is important)
        while not EOF():

            # double-quoted strings
            # embeds are doubled
            # end marks token
            if inString:
                if currentChar() == "\"":
                    if nextChar() == "\"":
                        token += "\""
                        offset += 1
                    else:
                        inString = False
                        tokens.add(token, self.TOK_TYPE_OPERAND, self.TOK_SUBTYPE_TEXT)
                        token = ""
                else:
                    token += currentChar()
                offset += 1
                continue

            # single-quoted strings (links)
            # embeds are double
            # end does not mark a token
            if inPath:
                if currentChar() == "'":
                    if nextChar() == "'":
                        token += "'"
                        offset += 1
                    else:
                        inPath = False
                else:
                    token += currentChar()
                offset += 1;
                continue;

                # bracketed strings (range offset or linked workbook name)
            # no embeds (changed to "()" by Excel)
            # end does not mark a token
            if inRange:
                if currentChar() == "]":
                    inRange = False
                token += currentChar()
                offset += 1
                continue

            # error values
            # end marks a token, determined from absolute list of values
            if inError:
                token += currentChar()
                offset += 1
                if ",#NULL!,#DIV/0!,#VALUE!,#REF!,#NAME?,#NUM!,#N/A,".find("," + token + ",") != -1:
                    inError = False
                    tokens.add(token, self.TOK_TYPE_OPERAND, self.TOK_SUBTYPE_ERROR)
                    token = ""
                continue;

            # independent character evaulation (order not important)
            #
            # establish state-dependent character evaluations
            if currentChar() == "\"":
                if len(token) > 0:
                    # not expected
                    tokens.add(token, self.TOK_TYPE_UNKNOWN)
                    token = ""
                inString = True
                offset += 1
                continue

            if currentChar() == "'":
                if len(token) > 0:
                    # not expected
                    tokens.add(token, self.TOK_TYPE_UNKNOWN)
                    token = ""
                inPath = True
                offset += 1
                continue

            if (currentChar() == "["):
                inRange = True
                token += currentChar()
                offset += 1
                continue

            if (currentChar() == "#"):
                if (len(token) > 0):
                    # not expected
                    tokens.add(token, self.TOK_TYPE_UNKNOWN)
                    token = ""
                inError = True
                token += currentChar()
                offset += 1
                continue

            # mark start and end of arrays and array rows
            if (currentChar() == "{"):
                if (len(token) > 0):
                    # not expected
                    tokens.add(token, self.TOK_TYPE_UNKNOWN)
                    token = ""
                tokenStack.push(tokens.add("ARRAY", self.TOK_TYPE_FUNCTION, self.TOK_SUBTYPE_START))
                tokenStack.push(tokens.add("ARRAYROW", self.TOK_TYPE_FUNCTION, self.TOK_SUBTYPE_START))
                offset += 1
                continue

            if (currentChar() == ";"):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.addRef(tokenStack.pop())
                tokens.add(",", self.TOK_TYPE_ARGUMENT)
                tokenStack.push(tokens.add("ARRAYROW", self.TOK_TYPE_FUNCTION, self.TOK_SUBTYPE_START))
                offset += 1
                continue

            if (currentChar() == "}"):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.addRef(tokenStack.pop())
                tokens.addRef(tokenStack.pop())
                offset += 1
                continue

            # trim white-space
            if (currentChar() == " "):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.add("", self.TOK_TYPE_WSPACE)
                offset += 1
                while ((currentChar() == " ") and (not EOF())):
                    offset += 1
                continue

            # multi-character comparators
            if (",>=,<=,<>,".find("," + doubleChar() + ",") != -1):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.add(doubleChar(), self.TOK_TYPE_OP_IN, self.TOK_SUBTYPE_LOGICAL)
                offset += 2
                continue

            # standard infix operators
            if ("+-*/^&=><".find(currentChar()) != -1):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.add(currentChar(), self.TOK_TYPE_OP_IN)
                offset += 1
                continue

            # standard postfix operators
            if ("%".find(currentChar()) != -1):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.add(currentChar(), self.TOK_TYPE_OP_POST)
                offset += 1
                continue

            # start subexpression or function
            if (currentChar() == "("):
                if (len(token) > 0):
                    tokenStack.push(tokens.add(token, self.TOK_TYPE_FUNCTION, self.TOK_SUBTYPE_START))
                    token = ""
                else:
                    tokenStack.push(tokens.add("", self.TOK_TYPE_SUBEXPR, self.TOK_SUBTYPE_START))
                offset += 1
                continue

            # function, subexpression, array parameters
            if (currentChar() == ","):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                if (not (tokenStack.type() == self.TOK_TYPE_FUNCTION)):
                    tokens.add(currentChar(), self.TOK_TYPE_OP_IN, self.TOK_SUBTYPE_UNION)
                else:
                    tokens.add(currentChar(), self.TOK_TYPE_ARGUMENT)
                offset += 1
                continue

            # stop subexpression
            if (currentChar() == ")"):
                if (len(token) > 0):
                    tokens.add(token, self.TOK_TYPE_OPERAND)
                    token = ""
                tokens.addRef(tokenStack.pop())
                offset += 1
                continue

            # token accumulation
            token += currentChar()
            offset += 1

        # dump remaining accumulation
        if (len(token) > 0):
            tokens.add(token, self.TOK_TYPE_OPERAND)

        # move all tokens to a new collection, excluding all unnecessary white-space tokens
        tokens2 = f_tokens()

        while (tokens.moveNext()):
            token = tokens.current();

            if (token.ttype == self.TOK_TYPE_WSPACE):
                if ((tokens.BOF()) or (tokens.EOF())):
                    pass
                elif (not (
                                ((tokens.previous().ttype == self.TOK_TYPE_FUNCTION) and (
                                            tokens.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                                ((tokens.previous().ttype == self.TOK_TYPE_SUBEXPR) and (
                                            tokens.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                            (tokens.previous().ttype == self.TOK_TYPE_OPERAND)
                )
                      ):
                    pass
                elif (not (
                                ((tokens.next().ttype == self.TOK_TYPE_FUNCTION) and (
                                            tokens.next().tsubtype == self.TOK_SUBTYPE_START)) or
                                ((tokens.next().ttype == self.TOK_TYPE_SUBEXPR) and (
                                            tokens.next().tsubtype == self.TOK_SUBTYPE_START)) or
                            (tokens.next().ttype == self.TOK_TYPE_OPERAND)
                )
                      ):
                    pass
                else:
                    tokens2.add(token.tvalue, self.TOK_TYPE_OP_IN, self.TOK_SUBTYPE_INTERSECT)
                continue

            tokens2.addRef(token);

        # switch infix "-" operator to prefix when appropriate, switch infix "+" operator to noop when appropriate, identify operand
        # and infix-operator subtypes, pull "@" from in front of function names
        while (tokens2.moveNext()):
            token = tokens2.current()
            if ((token.ttype == self.TOK_TYPE_OP_IN) and (token.tvalue == "-")):
                if (tokens2.BOF()):
                    token.ttype = self.TOK_TYPE_OP_PRE
                elif (
                                    ((tokens2.previous().ttype == self.TOK_TYPE_FUNCTION) and (
                                                tokens2.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                                    ((tokens2.previous().ttype == self.TOK_TYPE_SUBEXPR) and (
                                                tokens2.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                                (tokens2.previous().ttype == self.TOK_TYPE_OP_POST) or
                            (tokens2.previous().ttype == self.TOK_TYPE_OPERAND)
                ):
                    token.tsubtype = self.TOK_SUBTYPE_MATH;
                else:
                    token.ttype = self.TOK_TYPE_OP_PRE
                continue

            if ((token.ttype == self.TOK_TYPE_OP_IN) and (token.tvalue == "+")):
                if (tokens2.BOF()):
                    token.ttype = self.TOK_TYPE_NOOP
                elif (
                                    ((tokens2.previous().ttype == self.TOK_TYPE_FUNCTION) and (
                                                tokens2.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                                    ((tokens2.previous().ttype == self.TOK_TYPE_SUBEXPR) and (
                                                tokens2.previous().tsubtype == self.TOK_SUBTYPE_STOP)) or
                                (tokens2.previous().ttype == self.TOK_TYPE_OP_POST) or
                            (tokens2.previous().ttype == self.TOK_TYPE_OPERAND)
                ):
                    token.tsubtype = self.TOK_SUBTYPE_MATH
                else:
                    token.ttype = self.TOK_TYPE_NOOP
                continue

            if ((token.ttype == self.TOK_TYPE_OP_IN) and (len(token.tsubtype) == 0)):
                if (("<>=").find(token.tvalue[0:1]) != -1):
                    token.tsubtype = self.TOK_SUBTYPE_LOGICAL
                elif (token.tvalue == "&"):
                    token.tsubtype = self.TOK_SUBTYPE_CONCAT
                else:
                    token.tsubtype = self.TOK_SUBTYPE_MATH
                continue

            if ((token.ttype == self.TOK_TYPE_OPERAND) and (len(token.tsubtype) == 0)):
                try:
                    float(token.tvalue)
                except ValueError, e:
                    if ((token.tvalue == 'TRUE') or (token.tvalue == 'FALSE')):
                        token.tsubtype = self.TOK_SUBTYPE_LOGICAL
                    else:
                        token.tsubtype = self.TOK_SUBTYPE_RANGE
                else:
                    token.tsubtype = self.TOK_SUBTYPE_NUMBER
                continue

            if (token.ttype == self.TOK_TYPE_FUNCTION):
                if (token.tvalue[0:1] == "@"):
                    token.tvalue = token.tvalue[1:]
                continue

        tokens2.reset();

        # move all tokens to a new collection, excluding all noops
        tokens = f_tokens()
        while (tokens2.moveNext()):
            if (tokens2.current().ttype != self.TOK_TYPE_NOOP):
                tokens.addRef(tokens2.current())

        tokens.reset()
        return tokens

    def parse(self, formula):
        self.tokens = self.getTokens(formula)

    def render(self):
        output = ""
        if self.tokens:
            for t in self.tokens.items:
                if t.ttype == self.TOK_TYPE_FUNCTION and t.tsubtype == self.TOK_SUBTYPE_START:
                    output += t.tvalue + "("
                elif t.ttype == self.TOK_TYPE_FUNCTION and t.tsubtype == self.TOK_SUBTYPE_STOP:
                    output += ")"
                elif t.ttype == self.TOK_TYPE_SUBEXPR and t.tsubtype == self.TOK_SUBTYPE_START:
                    output += "("
                elif t.ttype == self.TOK_TYPE_SUBEXPR and t.tsubtype == self.TOK_SUBTYPE_STOP:
                    output += ")"
                # TODO: add in RE substitution of " with "" for strings
                elif t.ttype == self.TOK_TYPE_OPERAND and t.tsubtype == self.TOK_SUBTYPE_TEXT:
                    output += "\"" + t.tvalue + "\""
                elif t.ttype == self.TOK_TYPE_OP_IN and t.tsubtype == self.TOK_SUBTYPE_INTERSECT:
                    output += " "

                else:
                    output += t.tvalue
        return output

    def prettyprint(self):
        indent = 0
        output = ""
        if self.tokens:
            for t in self.tokens.items:

                if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                    indent -= 1

                output += "    " * indent + t.tvalue + " <" + t.ttype + "> <" + t.tsubtype + ">" + "\n"

                if (t.tsubtype == self.TOK_SUBTYPE_START):
                    indent += 1;
        return output

    def get_dic_depth_token(self):
        indent = 0
        output = ""
        isiffunction = False

        dic_depth_token = {}
        if self.tokens:
            for t in self.tokens.items:

                if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                    indent -= 1
                    isiffunction = False

                if dic_depth_token.has_key(indent):
                    dic_depth_token[indent].append(t)
                else:
                    dic_depth_token[indent] = [t]

                if (t.tsubtype == self.TOK_SUBTYPE_START):
                    indent += 1;

        return dic_depth_token

    def get_inner_if(self):
        returnstring = 'IF('
        indent = 0
        stopindent = 0

        count = 0
        stop = False
        self.tokens.reset()
        while (self.tokens.moveNext()):
            token = self.tokens.current();

            t = token

            if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                indent -= 1
                if indent == (stopindent-1):
                    returnstring += t.tvalue + ')'
                    stop = False
                    return returnstring

            if stop:

                if (t.tsubtype == self.TOK_SUBTYPE_START ):

                    returnstring += t.tvalue + '('
                elif (t.tsubtype == self.TOK_SUBTYPE_STOP):

                    returnstring += t.tvalue + ')'
                else:
                    if t.tsubtype == 'text':
                        returnstring += '\"'+t.tvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        returnstring += '\"\"'
                    else:
                        returnstring += t.tvalue


            if (t.tsubtype == self.TOK_SUBTYPE_START):
                indent += 1;
                if not stop:
                    if (t.ttype == self.TOK_TYPE_FUNCTION and t.tvalue == 'IF') and indent > 1:
                        stop = True
                        stopindent = indent
        return returnstring
    def get_para_if(self):
        returnstring = 'IF('
        indent = 0
        stopindent = 0

        count = 0
        stop = False
        self.tokens.reset()
        para_if_list = []
        while (self.tokens.moveNext()):

            token = self.tokens.current();

            t = token
            newvalue = t.tvalue

            # if '!' in t.tvalue and t.tsubtype == 'range':
            #     newvalue = '\'' + t.tvalue.split('!')[0] + '\'' + '!' + t.tvalue.split('!')[1]
            #     isrange_anothersheet = True


            if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                indent -= 1

                if indent == (stopindent-1) :
                    returnstring += t.tvalue + ')'
                    stop = False
                    para_if_list.append(returnstring)
                    returnstring = 'IF('



            if stop:

                if (t.tsubtype == self.TOK_SUBTYPE_START ):

                    returnstring += newvalue + '('
                elif (t.tsubtype == self.TOK_SUBTYPE_STOP):

                    returnstring += newvalue + ')'
                else:
                    if t.tsubtype == 'text':
                        returnstring += '\"'+newvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        returnstring += '\"\"'
                    else:
                        returnstring += newvalue





            if (t.tsubtype == self.TOK_SUBTYPE_START ):
                indent += 1;
                if not stop:
                    if (t.ttype == self.TOK_TYPE_FUNCTION and t.tvalue == 'IF'):
                        stop = True
                        stopindent = indent
        return para_if_list


    def get_para_if_second(self):
        returnstring = 'IF('
        indent = 0
        stopindent = 0

        count = 0
        stop = False
        self.tokens.reset()
        para_if_list = []
        while (self.tokens.moveNext()):


            token = self.tokens.current();

            t = token
            newvalue = t.tvalue




            if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                indent -= 1


                if indent == (stopindent-1):
                    returnstring += t.tvalue + ')'
                    stop = False

                    para_if_list.append(returnstring)
                    returnstring = 'IF('



            if stop:

                if (t.tsubtype == self.TOK_SUBTYPE_START):

                    returnstring += newvalue + '('
                elif (t.tsubtype == self.TOK_SUBTYPE_STOP):

                    returnstring += newvalue + ')'
                else:
                    if t.tsubtype == 'text':
                        returnstring += '\"'+newvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        returnstring += '\"\"'
                    else:
                        returnstring += newvalue





            if (t.tsubtype == self.TOK_SUBTYPE_START ):
                indent += 1;
                if not stop:
                    if (t.ttype == self.TOK_TYPE_FUNCTION and t.tvalue == 'IF'):

                        stop = True
                        stopindent = indent

        return para_if_list


    def get_all_innerif_list(self):
        depth = self.get_nested_ifs()
        all_inner_list = []

        for eachdepth in range(1, depth):

            thisdepthinner = self.get_inner_if()

            try:
                self.parse(thisdepthinner)
                all_inner_list.append(thisdepthinner)
            except:
                break

        return all_inner_list

    def get_threeparts_IF(self):
        count = 0
        condition_string = ''
        truevalue_string = ''
        falsevalue_string = ''
        indent = 0
        inner = False


        returnlist = []

        if self.tokens:
            for t in self.tokens.items:
                isrange_anothersheet = False
                newvalue = ''


                # if '!' in t.tvalue and t.tsubtype == 'range':
                #     newvalue = '\''+t.tvalue.split('!')[0]+'\''+'!'+t.tvalue.split('!')[1]
                #     isrange_anothersheet = True



                if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                    inner = False
                    indent -= 1
                    if count == 0:
                        if indent != 0:
                            condition_string += ')'
                    elif count == 1:
                        truevalue_string += ')'
                    else:
                        if indent != 0:
                            falsevalue_string += ')'
                if indent == 1 and t.tvalue == ',' and not inner:
                    count += 1

                    continue
                if count == 0:
                    if isrange_anothersheet:
                        condition_string += newvalue

                    elif t.tsubtype == 'text':
                        condition_string += '\"'+t.tvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        condition_string += '\"\"'
                    else:
                        condition_string += t.tvalue
                elif count == 1:
                    if isrange_anothersheet:
                        truevalue_string += newvalue
                    elif t.tsubtype == 'text':
                        truevalue_string += '\"'+t.tvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        truevalue_string += '\"\"'
                    else:
                        truevalue_string += t.tvalue
                else:
                    if isrange_anothersheet:
                        falsevalue_string += newvalue
                    elif t.tsubtype == 'text':
                        falsevalue_string += '\"'+t.tvalue+'\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        falsevalue_string += '\"\"'
                    else:
                        falsevalue_string += t.tvalue

                if (t.tsubtype == self.TOK_SUBTYPE_START):
                    if indent != 0:
                        inner = True

                    if count == 0:
                        if indent != 0:
                            condition_string += '('

                    elif count == 1:
                        truevalue_string += '('
                    else:
                        falsevalue_string += '('
                    indent += 1;
        if count == 1:
            truevalue_string = truevalue_string[:-1]

        returnlist.append(condition_string[2:])
        returnlist.append(truevalue_string)
        returnlist.append(falsevalue_string)


        return returnlist

    #remove the parts not in if function:
    #if(if(a=b,c,d)==a)..



    def get_onlyIFfunction(self):
        returnstring = 'IF('
        indent = 0
        stopindent = 0

        stop = False
        self.tokens.reset()

        while (self.tokens.moveNext()):
            token = self.tokens.current();

            t = token

            if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                indent -= 1

                if indent == (stopindent - 1):
                    returnstring += t.tvalue + ')'
                    stop = False
                    return returnstring

            if stop:

                if (t.tsubtype == self.TOK_SUBTYPE_START):

                    returnstring += t.tvalue + '('
                elif (t.tsubtype == self.TOK_SUBTYPE_STOP):

                    returnstring += t.tvalue + ')'
                else:
                    if t.tsubtype == 'text':
                        returnstring += '\"' + t.tvalue + '\"'
                    elif t.tsubtype == 'number' and t.tvalue == '':
                        returnstring += '\"\"'
                    else:
                        returnstring += t.tvalue

            if (t.tsubtype == self.TOK_SUBTYPE_START):
                indent += 1;
                if not stop:
                    if (t.ttype == self.TOK_TYPE_FUNCTION and t.tvalue == 'IF'):
                        stop = True
                        stopindent = indent
        return returnstring

    def get_type_threeparts_IF(self):
        count = 0
        condition_string = ''
        truevalue_string = ''
        falsevalue_string = ''
        indent = 0

        returnlist = []

        if self.tokens:
            for t in self.tokens.items:

                if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                    indent -= 1
                    if count == 0:
                        if indent != 0:
                            condition_string += ')'
                    elif count == 1:
                        truevalue_string += ')'
                    else:
                        if indent != 0:
                            falsevalue_string += ')'

                if 'indent == 1' and t.tvalue == ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += t.tvalue
                elif count == 1:
                    truevalue_string += t.tvalue
                else:
                    falsevalue_string += t.tvalue

                if (t.tsubtype == self.TOK_SUBTYPE_START):

                    if count == 0:
                        if indent != 0:
                            condition_string += '('

                    elif count == 1:
                        truevalue_string += '('
                    else:
                        falsevalue_string += '('
                    indent += 1;

        returnlist.append(condition_string[2:])
        returnlist.append(truevalue_string)
        returnlist.append(falsevalue_string)

        return returnlist

    def returnTrue_if_is_MAXMIN_pattern(self):
        all_innerif_list = self.get_all_innerif_list()
        if len(all_innerif_list) == 0:
            return False

        for eachone in all_innerif_list:

            if eachone.count('IF') != 1:
                continue

            threeparts = self.get_threeparts_IF()

            if threeparts[1] + '>' + threeparts[2] == threeparts[0] or (threeparts[1] + '>' + '('+threeparts[2]+')' == threeparts[0]):
                return True
            if threeparts[2] + '>=' + threeparts[1] == threeparts[0] or (threeparts[1] + '>=' +'('+ threeparts[2]+')' == threeparts[0]):
                return True
            if threeparts[1] + '<' + threeparts[2] == threeparts[0] or (threeparts[1] + '<' + '('+threeparts[2]+')' == threeparts[0]):
                return True
            if threeparts[2] + '<=' + threeparts[1] == threeparts[0] or (threeparts[1] + '<=' +'('+ threeparts[2]+')' == threeparts[0]):
                return True

        return False

    def returnTrue_if_IFS_pattern(self,dic):
        # print dic

        condition_list = []
        truevalue_list = []
        falsevalue_list = []
        truetype_list = []



        # dic = self.get_dic_depth_token()
        # print dic

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                value = i.tvalue
                subtype = i.tsubtype
                if value is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += value
                elif count == 1:
                    truevalue_string += value
                    truetype_list.append(subtype)
                else:
                    falsevalue_string += value

            condition_list.append(condition_string)
            truevalue_list.append(truevalue_string)
            falsevalue_list.append(falsevalue_string)
        # print condition_list

        if not (len(set(falsevalue_list[1:][:-1])) == 1):
            return False

        if len(set(truevalue_list[1:])) == 1:
            return False


        for i in condition_list[1:]:
            if ('>' not in i) and ('<' not in i):
                return False

        return True



    def returnTrue_if_is_OR_pattern(self,formula,dic):
        # print dic
        # dic = self.get_dic_depth_token()

        condition_list = []
        truevalue_list = []
        falsevalue_list = []

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                i = i.tvalue
                if i is ',':
                    count += 1
                    continue

                if count == 0:
                    condition_string += i
                elif count == 1:
                    truevalue_string += i
                else:
                    falsevalue_string += i

            truevalue_list.append(truevalue_string)

            falsevalue_list.append(falsevalue_string)

        if not (len(set(truevalue_list[1:])) == 1):
            return False
        if not (len(set(falsevalue_list[1:])) == 2):
            return False
        if not set(falsevalue_list[1:][:-1]) == set(['IF']):
            return False
        truepart = self.get_threeparts_IF()[1]
        if truepart != '':
            if not formula.count(','+truepart+',') == (len(falsevalue_list)-1):
                return False
        else:
            if not formula.count(',\"\",') == (len(falsevalue_list)-1):
                return False
        return True
#already a contain equal formula (###may be wrong)
    def returnTrue_if_is_CHOOSE_pattern(self,dic):
        condition_list = []
        truevalue_list = []
        falsevalue_list = []


        returnlist = []

        # print dic

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                value = i.tvalue
                subtype = i.tsubtype
                if value is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += value
                elif count == 1:
                    truevalue_string += value

                else:
                    falsevalue_string += value

            condition_list.append(condition_string)
            truevalue_list.append(truevalue_string)
            falsevalue_list.append(falsevalue_string)
        print condition_list[1:]
        print truevalue_list[1:]
        print falsevalue_list[1:]


    def returnTrue_if_is_AND_pattern(self,formula,dic):
        # print dic
        # dic = self.get_dic_depth_token()

        condition_list = []
        truevalue_list = []
        falsevalue_list = []

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                i = i.tvalue
                if i is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += i
                elif count == 1:
                    truevalue_string += i
                else:
                    falsevalue_string += i

            truevalue_list.append(truevalue_string)

            falsevalue_list.append(falsevalue_string)

        if not (len(set(truevalue_list[1:])) == 2):
            return False
        if not set(truevalue_list[1:][:-1]) == set(['IF']):
            return False
        if len(falsevalue_list) == 1:
            return False
        # falsepart = self.get_threeparts_IF()[2]
        # print falsepart
        #
        #
        # if not formula.count(',' + falsepart) == (len(truevalue_list)):
        #     return False
        if not len(set(falsevalue_list)) == 1:
            return False

        return True

    def returnTrue_if_contain_equal_pattern(self,dic):
        # print dic

        condition_list = []
        truevalue_list = []
        falsevalue_list = []
        truetype_list = []

        returnlist = []

        # print dic

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                value = i.tvalue
                subtype = i.tsubtype
                if value is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += value
                elif count == 1:
                    truevalue_string += value
                    truetype_list.append(subtype)
                else:
                    falsevalue_string += value

            condition_list.append(condition_string)
            truevalue_list.append(truevalue_string)
            falsevalue_list.append(falsevalue_string)
        # print condition_list

        if not (len(set(falsevalue_list[1:][:-1])) == 1):
            return False

        if len(set(truevalue_list[1:])) == 1:
            return False

        beforeequallist = []
        afterequallist = []
        for i in condition_list[1:]:
            if ('=' not in i) or ('>' in i) or ('<' in i):
                return False
            beforeequallist.append(i.split('=')[0])


            afterequallist.append(i.split('=')[1])

        if not (len(set(beforeequallist)) == 1):
            return False
        returnlist.append(afterequallist)
        returnlist.append(truevalue_list)
        return returnlist

    def returnTrue_if_is_LOOKUP_pattern(self,dic):
        # print dic

        condition_list = []
        truevalue_list = []
        falsevalue_list = []
        truetype_list = []

        returnlist = []


        # print dic

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                value = i.tvalue
                subtype = i.tsubtype
                if value is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += value
                elif count == 1:
                    truevalue_string += value
                    truetype_list.append(subtype)
                else:
                    falsevalue_string += value

            condition_list.append(condition_string)
            truevalue_list.append(truevalue_string)
            falsevalue_list.append(falsevalue_string)
        # print condition_list


        if not set(truetype_list) == set(['range']):
            return False

        if not (len(set(falsevalue_list[1:][:-1])) == 1):
            return False

        if len(set(truevalue_list[1:])) == 1:
            return False

        beforeequallist = []
        afterequallist = []
        for i in condition_list[1:]:
            if '=' not in i:
                return False
            beforeequallist.append(i.split('=')[0])

            try:
                float(i.split('=')[1])
                return False
            except:
                afterequallist.append(i.split('=')[1])

        if not (len(set(beforeequallist)) == 1):
            return False
        returnlist.append(afterequallist)
        returnlist.append(truevalue_list)
        return returnlist

    # get the depth of nested ifs
    def get_nested_ifs(self):
        indent = 0
        if_indent = 0
        ifdepth = 0
        output = ""
        if self.tokens:
            for t in self.tokens.items:


                if if_indent >= ifdepth:
                    ifdepth = if_indent

                if (t.tsubtype == self.TOK_SUBTYPE_STOP):
                    indent -= 1
                    if (t.ttype == self.TOK_TYPE_FUNCTION):
                        if_indent -= 1
                output = "    " * indent + t.tvalue + " <" + t.ttype + "> <" + t.tsubtype + ">" + "\n"


                # print t.tsubtype
                if (t.tsubtype == self.TOK_SUBTYPE_START):
                    indent += 1;
                    if (t.ttype == self.TOK_TYPE_FUNCTION and t.tvalue == 'IF'):
                        if_indent += 1
                        ifdepth += 1



        # print 'if indent: ', if_indent

        return ifdepth

    def prettyprinttest(self):
        indent = 0
        output = ""
        tvaluelist = []
        ttypelist = []
        tsubtypelist = []
        resultlist = []
        if self.tokens:
            for t in self.tokens.items:
                tvaluelist.append(t.tvalue)

                ttypelist.append(t.ttype)

                tsubtypelist.append(t.tsubtype)

        resultlist.append(tvaluelist)
        resultlist.append(ttypelist)
        resultlist.append(tsubtypelist)

        return resultlist

        # already a contain equal formula. dic[condition] = value
    def get_list_threeparts(self,formula, threepartlist):

        # print dic
        count = 0

        condition_list = []
        truevalue_list = []
        falsevalue_list = []

        condition_string = ''
        truevalue_string = ''
        falsevalue_string = ''

        try:
            self.parse('='+formula)
            formula = self.get_onlyIFfunction()

            self.parse('=' + formula)
        except:
            return threepartlist


        threeparts = self.get_threeparts_IF()
        # print '#############', threeparts




        if threeparts == ['', '', '']:
            return threepartlist
        partcount = 0
        for eachpart in threeparts:
            partcount+=1
            try:
                self.parse('='+eachpart)
            except:
                return threepartlist
            formula = self.get_onlyIFfunction()

            threepartlist[partcount%3-1].append(eachpart)

            if formula and 'IF' in eachpart:
                self.get_list_threeparts(eachpart, threepartlist)
            else:
                count += 1
                if count == 3:
                    return threepartlist
                continue




    def get_dic_for_equal(self, dic):
        condition_list = []
        truevalue_list = []
        falsevalue_list = []
        dic_condition_value = {}

        returnlist = []

        returnlist = []

        # print dic

        for eachlist in dic:

            depthone_list = dic[eachlist]
            count = 0
            condition_string = ''
            truevalue_string = ''
            falsevalue_string = ''

            for i in depthone_list:
                value = i.tvalue
                subtype = i.tsubtype
                if value is ',':
                    count += 1
                    continue
                if count == 0:
                    condition_string += value
                elif count == 1:
                    truevalue_string += value

                else:
                    falsevalue_string += value

            condition_list.append(condition_string)
            truevalue_list.append(truevalue_string)
            falsevalue_list.append(falsevalue_string)
        condition_list = condition_list[1:]
        truevalue_list =  truevalue_list[1:]
        falsevalue_list =  falsevalue_list[1:]

        ct = 0
        beforequallist = []
        for i in condition_list:
            condition = i.split('=')[1]
            beforequallist.append(i.split('=')[0])
            dic_condition_value[condition] = truevalue_list[ct]
            ct+=1
        if len(set(beforequallist)) != 1:
            return False

        returnlist.append(dic_condition_value)
        returnlist.append(falsevalue_list[-1])
        returnlist.append(condition_list)
        returnlist.append(truevalue_list)
        returnlist.append(falsevalue_list)
        return returnlist


    # ---------------------------------------------------------------------------------------------
    def get_path(self, pathstring):
        pathstring = pathstring.strip()
        basicpath = 'D:\\Users\\v-jizha4\\ExcelExp\\XLData2\\'
        excelname = pathstring.split('/')[-1]
        output = basicpath + pathstring.replace(excelname, '').replace('/', '\\')
        return output

    def get_functionlist(self, txtpathstring):
        readfile = open(txtpathstring)
        totallines = readfile.readlines()
        lineswithouttab = []
        for eachline in totallines:
            lineswithouttab.append(eachline.strip())
        return lineswithouttab

    def if_longeststring(self, string, list):
        maxlen = max(len(s) for s in list)
        if len(string) == maxlen:
            return True
        else:
            return False

    # get the dic. Key: depth value num: the number of formulas that have this depth
    def getDic_depth_num(self, eachtxtfilepath, previousdic):

        functionlist = self.get_functionlist(eachtxtfilepath)
        returnlist = []
        longformulalist = []

        for i in functionlist:
            try:

                self.parse(i)
                thisformudepth = self.get_nested_ifs()
                if thisformudepth > 1:
                    longformulalist.append(eachtxtfilepath)
                    longformulalist.append(i)

                if thisformudepth != 0:

                    if previousdic.has_key(thisformudepth):
                        previousdic[thisformudepth] += 1
                    else:
                        previousdic[thisformudepth] = 1
            except:
                continue

        returnlist.append(previousdic)
        returnlist.append(longformulalist)
        return returnlist

    # print the number of files with different depth
    def getAll_NestedIf_Depth(self, filenamestring, limitfirst, limitsecond):
        pathlistfile = 'D:\\Users\\v-jizha4\\ExcelExp\\XLData2.log\\' + filenamestring
        readfile = open(pathlistfile)
        allpathlist = readfile.readlines()
        partpathlist = allpathlist[limitfirst:limitsecond]

        dic_depth_formunum = {}

        resultfilepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\NestedIfFileAndFormula-10000.txt'
        writefile = open(resultfilepath, 'w')
        filewithformulacount = 0

        path_noduplication = []
        for eachpath in partpathlist:
            path = self.get_path(eachpath)
            if path not in path_noduplication:
                path_noduplication.append(path)

        for eachpathstring in path_noduplication:

            if not os.path.isdir(eachpathstring):
                continue

            alltxtfiles = os.listdir(eachpathstring)

            for eachtxtfile in alltxtfiles:

                eachtxtfilepath = eachpathstring + eachtxtfile
                # deal with each file
                if os.path.isfile(eachtxtfilepath):
                    filewithformulacount = filewithformulacount + 1
                    print "Num of files:  " + str(filewithformulacount)
                    resultslist = self.getDic_depth_num(eachtxtfilepath, dic_depth_formunum)
                    dic_depth_formunum = resultslist[0]
                    longformulist = resultslist[1]

                    for i in longformulist:
                        writefile.write(i + '\n')

                    print dic_depth_formunum

        writefile.close()

    def get_nestedIF(self, filenamestring):
        filepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\' + filenamestring
        readfile = open(filepath)
        lines = readfile.readlines()
        returnlist = []
        for eachline in lines:
            if 'D:\Users' in eachline:
                continue
            eachline = eachline.strip()
            returnlist.append(eachline)
        return returnlist

    def contains(self, small, big):
        for i in xrange(len(big) - len(small) + 1):

            for j in xrange(len(small)):

                if big[i + j] != small[j]:
                    break
            else:

                return [i, i + len(small) - 1]
        return False



    def shorten_lookup(self):
        returnlist = self.returnTrue_if_is_LOOKUP_pattern()
        afterequallist = returnlist[0]
        truevaluelist = returnlist[1]

        wb = load_workbook(filename='D:\\Users\\v-jizha4\\666.xlsx')
        sheet_ranges = wb['ddd']
        thislist = truevaluelist

        newlist = afterequallist
        # the row and column list of the assigned values
        columnlist = []
        rowlist = []
        sheetnamelist = []
        cellnamelist = []

        for each in thislist:
            if '!' in each:
                sheetname = each.split('!')[0]
                sheetnamelist.append(sheetname)
                cellname = each.split('!')[1]
                cellnamelist.append(cellname)
                ws = wb.get_sheet_by_name(sheetname)
                rowlist.append(ws[cellname].row)
                columnlist.append(ws[cellname].column)

        if len(set(sheetnamelist)) != 1:
            print 'not for lookup'
        if len(set(rowlist)) == 1:
            print 'maybe hlookup'

            for row in ws.iter_rows(row_offset=1):

                celllist = []
                goalwhichcolumnlist = []
                for cell in row:
                    rownum = cell.row
                    celllist.append(str(cell.value).replace('u', ''))
                    goalwhichcolumnlist.append(cell.column)

                iftrue = self.contains(newlist, celllist)
                goalwhichcolumnlist = goalwhichcolumnlist[iftrue[0]:iftrue[1] + 1]

                if iftrue:
                    if columnlist[0] == goalwhichcolumnlist[0] and columnlist[-1] == goalwhichcolumnlist[-1]:
                        print 'perfect'
                        print goalwhichcolumnlist[0] + str(rownum)
                        print goalwhichcolumnlist[-1] + str(rowlist[0])

                    break
        if len(set(columnlist)) == 1:
            print 'maybe vlookup'
            for column in ws.iter_columns(column_offset=1):
                celllist = []
                goalwhichrowlist = []

                for cell in column:
                    celllist.append(cell.value)
                    goalwhichrowlist.append(cell.row)

                iftrue = self.contains(newlist, celllist)
                goalwhichrowlist = goalwhichrowlist[iftrue[0]:iftrue[1] + 1]

                if iftrue:
                    print iftrue
                    if columnlist[0] == goalwhichrowlist[0] and columnlist[-1] == goalwhichrowlist[-1]:
                        print 'perfect'
                    break


                    # print get_iden_functions(functionlist)

def generate_lookup_file():
    p = ExcelParser()

    # p.getAll_NestedIf_Depth('done-round1.txt',0,49114)
    resultfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\lookupformulas.txt'
    writefile = open(resultfilename,'w')

    filenamestring = 'NestedIfFileAndFormula-10000.txt'
    filepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\' + filenamestring
    readfile = open(filepath)
    lines = readfile.readlines()
    returnlist = []
    filenamelist = []
    for eachline in lines:
        eachline = eachline.strip()
        if 'D:\Users' in eachline:
            filenamelist.append(eachline)
            continue

        returnlist.append(eachline)
    analyzelist = returnlist
    count = 0
    lookupcount = 0

    for i in analyzelist:
        count+=1
        p.parse('='+i)
        dic = p.get_dic_depth_token()
        if p.returnTrue_if_is_LOOKUP_pattern():
            print '-----------------------'
            print filenamelist[count-1]
            print i
            lookupcount+=1
            print "lookup count: ",lookupcount
            print "total count: ",count
            writefile.write(filenamelist[count-1]+'\n')
            writefile.write(i+'\n')

    writefile.close()


def get_para_iflist_second(formula):
    # input: the formula that you want to refactor
    # output: a list of inner if parts. They do not contain each other. Return false if the formula is not nested if formula
    p = ExcelParser()
    try:
        p.parse('=' + formula)
    except:
        return False
    ifreturnfalse = False

    returnlist = p.get_para_if_second()


    if len(returnlist) == 0:

        return False

    newreturnlist = []




    for each in returnlist:

        if (each.count('IF') - each.count('IFS')) <= 1 and (each.count('IF') - each.count('IFERROR')) <=1 and (
            each.count('IF') - each.count('IFNA')) <= 1 and (each.count('IFS')!=0) and (each.count('IFERROR')!=0) and (each.count('IFNA')!=0):
            continue
        newreturnlist.append(each)

    return newreturnlist

def generate_containequal_file():
    p = ExcelParser()

    # p.getAll_NestedIf_Depth('done-round1.txt',0,49114)
    resultfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\containequalformulas.txt'
    writefile = open(resultfilename,'w')

    filenamestring = 'NestedIfFileAndFormula-10000.txt'
    filepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\' + filenamestring
    readfile = open(filepath)
    lines = readfile.readlines()
    returnlist = []
    filenamelist = []
    for eachline in lines:
        eachline = eachline.strip()
        if 'D:\Users' in eachline:
            filenamelist.append(eachline)
            continue

        returnlist.append(eachline)
    analyzelist = returnlist
    count = 0
    lookupcount = 0

    for i in analyzelist:
        count+=1
        p.parse('='+i)
        dic = p.get_dic_depth_token()
        if p.returnTrue_if_contain_equal_pattern():
            print '-----------------------'
            print filenamelist[count-1]
            print i
            lookupcount+=1
            print "contain equal count: ",lookupcount
            print "total count: ",count
            writefile.write(filenamelist[count-1]+'\n')
            writefile.write(i+'\n')

    writefile.close()


def generate_AND():
    p = ExcelParser()

    filenamestring = 'OTHER.txt'
    filepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\' + filenamestring
    readfile = open(filepath)

    andfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\andnew.txt'
    andwrite = open(andfilename,'w')

    orfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\ornew.txt'
    orwrite = open(orfilename, 'w')

    otherfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\othernew.txt'
    otherwrite = open(otherfilename, 'w')



    lines = readfile.readlines()
    returnlist = []
    filenamelist = []
    for eachline in lines:
        eachline = eachline.strip()
        if 'D:\Users' in eachline:
            filenamelist.append(eachline)
            continue

        returnlist.append(eachline)
    analyzelist = returnlist
    count = 0
    and_count = 0
    or_count = 0


    for i in analyzelist:
        count+=1
        p.parse('='+i)
        dic = p.get_dic_depth_token()
        if p.returnTrue_if_is_AND_pattern(i,dic):
            print '-----------------------'
            print filenamelist[count-1]
            print i
            and_count+=1
            print "AND count: ",and_count
            print "total count: ",count

            andwrite.write(filenamelist[count-1]+'\n')
            andwrite.write(i+'\n')
        elif p.returnTrue_if_is_OR_pattern(i, dic):
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            or_count += 1
            print "OR count: ", or_count
            print "total count: ", count
            orwrite.write(filenamelist[count - 1] + '\n')
            orwrite.write(i + '\n')
        else:

            otherwrite.write(filenamelist[count - 1] + '\n')
            otherwrite.write(i + '\n')

    andwrite.close()
    orwrite.close()
    otherwrite.close()




def generate():
    p = ExcelParser()

    filenamestring = 'NestedIfFileAndFormula-total.txt'
    filepath = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\' + filenamestring
    readfile = open(filepath)

    andfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\AND.txt'
    andwrite = open(andfilename,'w')
    orfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\OR.txt'
    orwrite = open(orfilename, 'w')
    lookupfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\LOOKUP.txt'
    lookupwrite = open(lookupfilename, 'w')
    conequalfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\CONEQUAL.txt'
    conequalwrite = open(conequalfilename, 'w')
    ifsfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\IFS.txt'
    ifswrite = open(ifsfilename, 'w')
    maxminfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\MAXMIN.txt'
    maxminwrite = open(maxminfilename, 'w')
    otherfilename = 'D:\\Users\\v-jizha4\\ExcelExp\\expAnalyResults\\totalsplits\\OTHER.txt'
    otherwrite = open(otherfilename, 'w')


    lines = readfile.readlines()
    returnlist = []
    filenamelist = []
    for eachline in lines:
        eachline = eachline.strip()
        if 'D:\Users' in eachline:
            filenamelist.append(eachline)
            continue

        returnlist.append(eachline)
    analyzelist = returnlist
    count = 0
    and_count = 0
    or_count = 0
    lookup_count = 0
    conequal_count = 0
    ifs_count = 0
    maxmin_count = 0

    for i in analyzelist:
        count+=1
        p.parse('='+i)
        dic = p.get_dic_depth_token()
        if p.returnTrue_if_is_AND_pattern(i,dic):
            print '-----------------------'
            print filenamelist[count-1]
            print i
            and_count+=1
            print "AND count: ",and_count
            print "total count: ",count
            andwrite.write(filenamelist[count-1]+'\n')
            andwrite.write(i+'\n')
        elif p.returnTrue_if_is_OR_pattern(i,dic):
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            or_count += 1
            print "OR count: ", or_count
            print "total count: ", count
            orwrite.write(filenamelist[count - 1] + '\n')
            orwrite.write(i + '\n')
        elif p.returnTrue_if_is_LOOKUP_pattern(dic):
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            lookup_count += 1
            print "LOOKUP count: ", lookup_count
            print "total count: ", count
            lookupwrite.write(filenamelist[count - 1] + '\n')
            lookupwrite.write(i+'\n')
        elif p.returnTrue_if_contain_equal_pattern(dic):
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            conequal_count += 1
            print "CON Equal count: ", conequal_count
            print "total count: ", count
            conequalwrite.write(filenamelist[count - 1] + '\n')
            conequalwrite.write(i+'\n')
        elif p.returnTrue_if_IFS_pattern(dic):
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            ifs_count += 1
            print "IFS count: ", ifs_count
            print "total count: ", count
            ifswrite.write(filenamelist[count - 1] + '\n')
            ifswrite.write(i+'\n')
        elif p.returnTrue_if_is_MAXMIN_pattern():
            print '-----------------------'
            print filenamelist[count - 1]
            print i
            maxmin_count += 1
            print "CON Equal count: ", maxmin_count
            print "total count: ", count
            maxminwrite.write(filenamelist[count - 1] + '\n')
            maxminwrite.write(i + '\n')
        else:
            otherwrite.write(filenamelist[count - 1] + '\n')
            otherwrite.write(i+'\n')



    andwrite.close()
    orwrite.close()
    lookupwrite.close()
    conequalwrite.close()
    ifswrite.close()
    maxminwrite.close()
    otherwrite.close()
# ========================================================================
# Main code:
#
# A simple test-rig.  Iterate through a list of test input strings,
# outputing a nested display of the token stream parsed from each one.
# ========================================================================
if __name__ == "__main__":

    formula = 'SUM(IF((DelPoint="4C")*IF((DType="firm")+(DType="econ")>0,1,0)*(OFFSET(DelPoint,0,RC+2)<0),OFFSET(DelPoint,0,RC+2),0))'
    p = ExcelParser()
    p.parse('='+formula)

    print p.get_nested_ifs()

    # print p.prettyprint()


